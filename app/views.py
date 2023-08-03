import pandas as pd
import pybliometrics, scholarly, openpyxl

from datetime import datetime
from .forms import *
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView
from django.views.generic.base import View
from django.db.models.functions import Concat
from django.db.models import F, Value
from .models import Scientist
from openpyxl.styles import Alignment, Font, Border
from pybliometrics.scopus import *
from wos import *
from scholarly import scholarly



class MainPage(View):
    """Отримання даних вчених для побудови графіку"""

    def get(self, request):
        google_scholar_h = Scientist.objects.all().order_by('-h_index_google_scholar', 'lastname_uk').filter(h_index_google_scholar__isnull=False, draft=False)[:10]
        scopus_h = Scientist.objects.all().order_by('-h_index_scopus', 'lastname_uk').filter(h_index_scopus__isnull=False, draft=False)[:10]
        publons_h = Scientist.objects.all().order_by('-h_index_publons', 'lastname_uk').filter(h_index_publons__isnull=False, draft=False)[:10]
        publons = Scientist.objects.all().order_by('-publons_count_pub', 'lastname_uk').filter(publons_count_pub__isnull=False, draft=False)[:10]
        scopus = Scientist.objects.all().order_by('-scopus_count_pub', 'lastname_uk').filter(scopus_count_pub__isnull=False, draft=False)[:10]

        # Publons Data
        scientist_publons = []
        publons_count_pub = []
        publons_profile_id = []

        for i in range(10):
            publons_name = (publons[i].lastname_uk + " " + publons[i].firstname_uk + " " + publons[i].middlename_uk)
            scientist_publons.append(publons_name)
            publons_pub = publons[i].publons_count_pub
            publons_count_pub.append(publons_pub)
            publons_id = publons[i].profile_id
            publons_profile_id.append(publons_id)

        # Scopus Data
        scientist_scopus = []
        scopus_count_pub = []
        scopus_profile_id = []

        for i in range(10):
            scopus_name = (scopus[i].lastname_uk + " " + scopus[i].firstname_uk + " " + scopus[i].middlename_uk)
            scientist_scopus.append(scopus_name)
            scopus_pub = scopus[i].scopus_count_pub
            scopus_count_pub.append(scopus_pub)
            scopus_id = scopus[i].profile_id
            scopus_profile_id.append(scopus_id)

        # Функція для будови context
        def generate_key(prefix, index):
            return f"{prefix}_{index + 1}"

        context = {}

        # Publons and Scopus Publications Top
        for i in range(10):
            context[generate_key('scientist_publons', i)] = scientist_publons[i]
            context[generate_key('scientist_scopus', i)] = scientist_scopus[i]
            context[generate_key('publons_count_pub', i)] = publons_count_pub[i]
            context[generate_key('scopus_count_pub', i)] = scopus_count_pub[i]
            context[generate_key('publons_profile_id', i)] = publons_profile_id[i]
            context[generate_key('scopus_profile_id', i)] = scopus_profile_id[i]

            # Scopus and Publons h index
            context[generate_key('scientist_scopus_h', i)] = f"{scopus_h[i].lastname_uk} {scopus_h[i].firstname_uk} {scopus_h[i].middlename_uk}"
            context[generate_key('scientist_publons_h', i)] = f"{publons_h[i].lastname_uk} {publons_h[i].firstname_uk} {publons_h[i].middlename_uk}"
            context[generate_key('scopus_h', i)] = scopus_h[i].h_index_scopus
            context[generate_key('publons_h', i)] = publons_h[i].h_index_publons
            context[generate_key('scopus_profile_id_h', i)] = scopus_h[i].profile_id
            context[generate_key('publons_profile_id_h', i)] = publons_h[i].profile_id

            # Google Scholar h index
            context[generate_key('scientist_google_scholar_h', i)] = f"{google_scholar_h[i].lastname_uk} {google_scholar_h[i].firstname_uk} {google_scholar_h[i].middlename_uk}"
            context[generate_key('google_scholar_h', i)] = google_scholar_h[i].h_index_google_scholar
            context[generate_key('google_scholar_profile_id_h', i)] = google_scholar_h[i].profile_id

        return render(request, "mainPage.html", context)


class ScientistsPage(ListView):
    """Перелік вчених"""

    model = Scientist
    queryset = Scientist.objects.order_by('lastname_uk').filter(draft=False)
    context_object_name = "scientist_list"
    template_name = "scientistsPage.html"
    paginate_by = 20

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['scientists_counter'] = self.queryset.count()
        context['form'] = SelectForm(initial={
            'select': self.request.GET.get('select', ''),
        })

        return context


class Search(ListView):
    """Пошук вчених та сортування за ознаками"""

    context_object_name = "scientist_list"
    paginate_by = 25
    template_name = "scientistsPage.html"

    def get_queryset(self):
        queryset = Scientist.objects.filter(draft=False)
        q = self.request.GET.get("q")
        select = self.request.GET.get('select')
        filter_dropdown_menu = self.request.GET.get('filter')

        filter_mapping = {
            'fullname_up': 'lastname_uk',
            'fullname_down': '-lastname_uk',
            'gsh_up': 'h_index_google_scholar',
            'gsh_down': '-h_index_google_scholar',
            'ph_up': 'h_index_publons',
            'ph_down': '-h_index_publons',
            'sh_up': 'h_index_scopus',
            'sh_down': '-h_index_scopus',
        }

        if "fullname" in select:
            queryset = queryset.annotate(
                fullname_uk=Concat(F('lastname_uk'), Value(' '), F('firstname_uk'), Value(' '), F('middlename_uk'))
            ).filter(fullname_uk__icontains=q).order_by(filter_mapping.get(filter_dropdown_menu, 'lastname_uk'))
        else:
            sort_field = filter_mapping.get(filter_dropdown_menu, 'lastname_uk')
            queryset = queryset.filter(department__title_department__icontains=q).order_by(sort_field)

        return queryset

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['scientists_counter'] = self.get_queryset().count()
        context['form'] = SelectForm(initial={
            'select': self.request.GET.get('select', ''),
        })
        return context


@login_required(login_url='/accounts/login/')
def update_scientists_records(request):
    error_log = ''
    scientists = Scientist.objects.all().filter(draft=False).order_by('date_update')
    for scientist in scientists:
        print(f"\n{scientist.profile_id} Profile {scientist.lastname_uk} {scientist.firstname_uk} {scientist.middlename_uk} updating!!!")
        print(f"Last date updating {scientist.date_update}")
        if scientist.scopusid:
            try:
                scopus_author = AuthorRetrieval(scientist.scopusid)

                # Зміненя даних у БД h-index, кількості публікацій
                scientist.h_index_scopus = scopus_author.h_index
                scientist.scopus_count_pub = scopus_author.document_count
                scientist.save()
                print(f"Scopus: SUCCESS!")

            except Exception as e:
                print(f"Scopus: ERROR!")
                error_log += f"\nScopus: Error updating data for {scientist.lastname_uk} {scientist.firstname_uk} https://s2m.ontu.edu.ua/profile/{scientist.profile_id}: {e}"
        else:
            print(f"Scopus: EMPTY!")
            
        if scientist.google_scholar:
            try:
                author = scholarly.search_author_id(f'{scientist.google_scholar}')
                author = (scholarly.fill(author, sections=['indices', 'publications']))
                
                scientist.h_index_google_scholar = author['hindex']
                scientist.google_scholar_count_pub = len(author['publications'])
                scientist.save()
                print(f"Google Scholar: SUCCESS!")


            except Exception as e:
                print(f"Google Scholar: ERROR!")
                error_log += f"Google Scholar: ERROR updating data for {scientist.lastname_uk} {scientist.firstname_uk} https://s2m.ontu.edu.ua/profile/{scientist.profile_id}: {e}"
        else:
            print(f"Google Scholar: EMPTY!")

    return HttpResponse("Update completed successfully.\n" + error_log)


class ProfilePage(View):
    """Сторінка вченого з усіма його даними"""

    def get(self, request, profile_id):
        profile_scientist = Scientist.objects.get(profile_id=profile_id)
        context = {
            'scientist': profile_scientist,
        }
        return render(request, 'profilePage.html', context)

    def post(self, request, profile_id):
        #api_scpopus_key = '5e5dc847f87e9db2f294456be2b932e4'

        # Web of Science

        # wos_api_key = 'your_web_of_science_api_key'
        # wos_id = Scientist.objects.get(profile_id=profile_id).publons
        # if wos_id:
        #     # wos_client = WosClient()

        #     # author_info = wos_client.retrieve_author(wos_id)

        #     # h_index = author_info.get('h_index', 0)
        #     # publication_count = author_info.get('document_count', 0)

        #     # # Обновляем поля в базе данных Django
        #     # scientist = Scientist.objects.get(profile_id=profile_id)
        #     # scientist.h_index_wos = h_index
        #     # scientist.publication_count_wos = publication_count
        #     # scientist.save()
        #     client = WosClient(wos_api_key)

        #     # Получаем данные о публикациях автора по его WOS ID
        #     query = f"AU_ID({wos_id})"
        #     response = client.search(query)

        #     if response.get('records_found', 0) > 0:
        #         record = response['records'][0]

        #         h_index = record.get('h_index', 0)
        #         publication_count = record.get('document_count', 0)

        #         scientist = Scientist.objects.get(profile_id=profile_id)
        #         scientist.h_index_publons = h_index
        #         scientist.publons_count_pub = publication_count
        #         scientist.save()

        # Scopus
        # Отримання необхідного scopusid за profile_id
        profile = Scientist.objects.get(profile_id=profile_id)
        if profile.scopusid:
            try:
                scopus_author = AuthorRetrieval(profile.scopusid)

                # Зміненя даних у БД та збереження
                profile.h_index_scopus = scopus_author.h_index
                profile.scopus_count_pub = scopus_author.document_count
                profile.save()
            except Exception as e:
                error_log = f"\nScopus: Error updating data for {profile.lastname_uk} {profile.firstname_uk} {profile.middlename_uk} https://s2m.ontu.edu.ua/profile/{scientist.profile_id}: {e}"
                return HttpResponse("Error Update.\n" + error_log)


        #Google Scholar
        
                
        if profile.google_scholar:
            try:
                author = scholarly.search_author_id(f'{profile.google_scholar}')
                author = (scholarly.fill(author, sections=['indices', 'publications']))
                
                profile.h_index_google_scholar = author['hindex']
                profile.google_scholar_count_pub = len(author['publications'])
                profile.save()
            
            except Exception as e:
                error_log = f"\Google Scholar: Error updating data for {profile.lastname_uk} {profile.firstname_uk} {profile.middlename_uk} https://s2m.ontu.edu.ua/profile/{scientist.profile_id}: {e}"
                return HttpResponse("Error Update.\n" + error_log)

        return redirect('profile', profile_id=profile_id)

        

def information(request):
    """Сторінка довідки"""
    return render(request, 'informationPage.html')


def report(request):
    """Сторінка звітів"""
    return render(request, 'reportPage.html')


@login_required(login_url='/accounts/login/')
def export_xlsx(request):
    '''Генерація файлу .xlsx з данми кожного вченого окрім тих які позначені як draft'''

    # Створюємо  Workbook
    workbook = openpyxl.Workbook()

    # Створюємо лист (worksheet)
    worksheet = workbook.active
    worksheet.title = "Звіт"

    queryset = Scientist.objects.filter(draft=False).prefetch_related(
        'department__faculty').order_by('lastname_uk')

    # Створюємо порожній DataFrame для зберігання даних
    data = pd.DataFrame(columns=['Факультет (Інститут)', 'Кафедра, відділ тощо', 'ПІБ', 'ID Scopus', 'Індекс Гірша Scopus',
                                 'Кількість публікацій Scopus', 'ID Web of Science',
                                 'Індекс Гірша Web of Science', 'Кількість публікацій WoS', 'Публікації Scopus', 'Публікації WoS'])

    # Заповнюємо DataFrame даними із queryset
    for scientist in queryset:

        full_name = ' '.join(
            [scientist.lastname_uk, scientist.firstname_uk, scientist.middlename_uk])
        title_department = scientist.department.title_department if scientist.department else ''
        title_faculty = scientist.department.faculty.title_faculty if scientist.department.faculty else ''

        # Отримуємо всі назви публікацій та об'єднуємо їх в один рядок
        scopus_publication_titles = ',\n'.join(
            [f"{index+1}. {publication.publication_title}" for index, publication in enumerate(scientist.publication_wos.all())])
        wos_publication_titles = ',\n'.join(
            [f"{index+1}. {publication.publication_title}" for index, publication in enumerate(scientist.publication_scopus.all())])

        # Додаємо дані в DataFrame
        data.loc[len(data)] = [title_faculty, title_department, full_name, scientist.scopusid, scientist.h_index_scopus,
                               scientist.scopus_count_pub, scientist.publons, scientist.h_index_publons,
                               scientist.publons_count_pub, scopus_publication_titles, wos_publication_titles]

    # worksheet.column_dimensions.group('A', 'B', hidden=True, outline_level=0)
    worksheet.column_dimensions['A'].width = 40
    worksheet.column_dimensions['B'].width = 35

    # worksheet.column_dimensions.group('C', 'D', hidden=True, outline_level=0)
    worksheet.column_dimensions['C'].width = 40
    worksheet.column_dimensions['D'].width = 18

    # worksheet.column_dimensions.group('E', 'I', hidden=True, outline_level=0)
    worksheet.column_dimensions['E'].width = 18
    worksheet.column_dimensions['F'].width = 18
    worksheet.column_dimensions['G'].width = 18
    worksheet.column_dimensions['H'].width = 18
    worksheet.column_dimensions['I'].width = 18

    # worksheet.column_dimensions.group('J', 'K', hidden=True, outline_level=0)
    worksheet.column_dimensions['J'].width = 85
    worksheet.column_dimensions['K'].width = 85

    # Форматування комірок
    border = Border(left=openpyxl.styles.Side(border_style='thin'), right=openpyxl.styles.Side(
        border_style='thin'), top=openpyxl.styles.Side(border_style='thin'), bottom=openpyxl.styles.Side(border_style='thin'))
    alignment = Alignment(horizontal='center',
                          vertical='center', wrap_text=True)
    font = Font(size=12)

    # Встановлення висоти комірок
    endRow = 800
    for i in range(0, endRow):
        worksheet.row_dimensions[i+1].height = 58

    # Заповнюємо таблицю Excel даними з DataFrame
    start_row_index = 2
    for row in data.itertuples(index=False):
        row_index = start_row_index
        for col_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = value

        start_row_index += 1

    # Застосовуємо форматування для кожного стовпця
    for column in worksheet.columns:
        for cell in column:
            cell.border = border
            cell.alignment = alignment
            cell.font = font

    headers = ['Факультет (Інститут)', 'Кафедра, відділ тощо', 'ПІБ', 'ID Scopus', 'Індекс Гірша Scopus',
               'Кількість публікацій Scopus', 'ID Web of Science',
               'Індекс Гірша Web of Science', 'Кількість публікацій WoS', 'Публікації Scopus', 'Публікації WoS']

    # Застосовуємо форматування для першого рядка
    row_index = 1
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row_index, column=col_index)
        cell.value = header
        cell.font = Font(bold=True, size=13)

    for cell in worksheet['D'][1:]:
        cell.font = Font(underline='single', color='0563C1')
        cell.hyperlink = f"https://www.scopus.com/authid/detail.uri?authorId={cell.value}"

    for cell in worksheet['G'][1:]:
        cell.font = Font(underline='single', color='0563C1')
        cell.hyperlink = f"https://publons.com/researcher/{cell.value}"

    filename = "Zvit" + datetime.now().strftime("_%d_%m_%Y") + ".xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    workbook.save(response)

    return response


@login_required(login_url='/accounts/login/')
def naukometria_xlsx(request):
    '''Генерація файлу .xlsx з данми кожного вченого окрім тих які позначені як draft'''

    # Створюємо  Workbook
    workbook = openpyxl.Workbook()

    # Створюємо лист (worksheet)
    worksheet = workbook.active
    worksheet.title = "Наукометрія"

    queryset = Scientist.objects.filter(draft=False).prefetch_related(
        'department__faculty__institute').order_by('lastname_uk')

    # Створюємо порожній DataFrame для зберігання даних
    data = pd.DataFrame(columns=['Інститут', 'Кафедра', 'ПІБ', 'Штат', 'Ступінь', 'Посади', 'Google Scholar', 'GS h-index', 'Кількість публікацій GS',
                        'ORCID', 'ID Scopus', 'Scopus h-index', 'Кількість публікацій Scopus', 'Researcher ID', 'WoS h-index', 'Кількість публікацій WoS'])

    # Заповнюємо DataFrame даними із queryset
    for scientist in queryset:

        full_name = ' '.join(
            [scientist.lastname_uk, scientist.firstname_uk, scientist.middlename_uk, scientist.profile_id])
        title_department = scientist.department.title_department if scientist.department else ''
        title_institute = scientist.department.faculty.institute.abbreviation if scientist.department.faculty.institute else ''
        title_degree = scientist.degree.title_degree if scientist.degree else ''
        title_post = scientist.post.title_post if scientist.post else ''

        if len(data) > 0:
            data.loc[len(data)] = [title_institute, title_department, full_name, scientist.staff, title_degree, title_post, scientist.google_scholar, scientist.h_index_google_scholar, scientist.google_scholar_count_pub, scientist.orcid, scientist.scopusid, scientist.h_index_scopus,
                                   scientist.scopus_count_pub, scientist.publons, scientist.h_index_publons,
                                   scientist.publons_count_pub]
        else:
            data.loc[0] = [title_institute, title_department, full_name, scientist.staff, title_degree, title_post, scientist.google_scholar, scientist.h_index_google_scholar, scientist.google_scholar_count_pub, scientist.orcid, scientist.scopusid, scientist.h_index_scopus,
                           scientist.scopus_count_pub, scientist.publons, scientist.h_index_publons,
                           scientist.publons_count_pub]

    # worksheet.column_dimensions.group('A', 'B', hidden=True, outline_level=0)
    worksheet.column_dimensions['A'].width = 18
    worksheet.column_dimensions['B'].width = 18

    # worksheet.column_dimensions.group('C', 'D', hidden=True, outline_level=0)
    worksheet.column_dimensions['C'].width = 40
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 14

    # worksheet.column_dimensions.group('E', 'I', hidden=True, outline_level=0)
    worksheet.column_dimensions['F'].width = 12
    worksheet.column_dimensions['G'].width = 10
    worksheet.column_dimensions['H'].width = 10
    worksheet.column_dimensions['I'].width = 16
    worksheet.column_dimensions['J'].width = 22
    worksheet.column_dimensions['K'].width = 16

    # worksheet.column_dimensions.group('J', 'K', hidden=True, outline_level=0)
    worksheet.column_dimensions['L'].width = 12
    worksheet.column_dimensions['M'].width = 16
    worksheet.column_dimensions['N'].width = 16
    worksheet.column_dimensions['O'].width = 10
    worksheet.column_dimensions['P'].width = 16

    # Форматування комірок
    border = Border(left=openpyxl.styles.Side(border_style='thin'), right=openpyxl.styles.Side(
        border_style='thin'), top=openpyxl.styles.Side(border_style='thin'), bottom=openpyxl.styles.Side(border_style='thin'))
    alignment = Alignment(horizontal='center',
                          vertical='center', wrap_text=True)
    font = Font(size=12)

    # Встановлення висоти комірок
    endRow = 800
    for i in range(0, endRow):
        worksheet.row_dimensions[i+1].height = 58

    # Заповнюємо таблицю Excel даними з DataFrame
    start_row_index = 2
    for row in data.itertuples(index=False):
        row_index = start_row_index
        for col_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = value

        start_row_index += 1

    # Застосовуємо форматування для кожного стовпця
    for column in worksheet.columns:
        for cell in column:
            cell.border = border
            cell.alignment = alignment
            cell.font = font

    headers = ['Інститут', 'Кафедра', 'ПІБ', 'Штат', 'Ступінь', 'Посади', 'Google Scholar', 'GS h-index', 'Кількість публікацій GS', 'ORCID',
               'ID Scopus', 'Scopus h-index', 'Кількість публікацій Scopus', 'Researcher ID', 'WoS h-index', 'Кількість публікацій WoS']

    # Застосовуємо форматування для першого рядка
    row_index = 1
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row_index, column=col_index)
        cell.value = header
        cell.font = Font(bold=True, size=12)

    for cell in worksheet['A'][1:]:
        cell.font = Font(size=11)

    for cell in worksheet['B'][1:]:
        cell.font = Font(size=10)

    for cell in worksheet['C'][1:]:
        cell.font = Font(color='5c8fad')
        cell.hyperlink = f"https://s2m.ontu.edu.ua/profile/{cell.value[-4:]}"
        cell.value = cell.value[:-5]
    
    for cell in worksheet['D'][1:]:
        if cell.value == 1:
            cell.value = 'Так'
        else:
            cell.value = 'Ні'
        cell.font = Font(size=10)
        
    
    for cell in worksheet['E'][1:]:
        cell.font = Font(size=10)

    for cell in worksheet['F'][1:]:
        cell.font = Font(size=10)

    for cell in worksheet['G'][1:]:
        cell.font = Font(underline='single', color='0563C1')
        cell.hyperlink = f"https://scholar.google.com/citations?user={cell.value}"
        if cell.value:
            cell.value = 'GS Link'

    for cell in worksheet['J'][1:]:
        cell.font = Font(size=11, underline='single', color='0563C1')
        cell.hyperlink = f"https://orcid.org/{cell.value}"

    for cell in worksheet['K'][1:]:
        cell.font = Font(underline='single', color='0563C1')
        cell.hyperlink = f"https://www.scopus.com/authid/detail.uri?authorId={cell.value}"

    for cell in worksheet['N'][1:]:
        cell.font = Font(underline='single', color='0563C1')
        cell.hyperlink = f"https://publons.com/researcher/{cell.value}"

    filename = "Scientometrics" + datetime.now().strftime("_%d_%m_%Y") + ".xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    workbook.save(response)

    return response
